import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font

INPUT_FILE = os.path.expanduser(
    "~/.local/share/opencode/tool-output/tool_ea7b0203f001ay6ur6BmWeqvE3"
)
OUTPUT_FILE = "/Users/somit17/PycharmProjects/CodeForces/problems_by_difficulty.xlsx"

with open(INPUT_FILE, "r") as f:
    data = json.load(f)

problems = data["result"]["problems"]
statistics = data["result"]["problemStatistics"]

stats_map = {}
for s in statistics:
    stats_map[(s["contestId"], s["index"])] = s["solvedCount"]

merged = []
for p in problems:
    if p.get("type") != "PROGRAMMING" or p.get("rating") is None:
        continue
    key = (p["contestId"], p["index"])
    solved = stats_map.get(key, 0)
    merged.append({
        "rating": p["rating"],
        "contestId": p["contestId"],
        "index": p["index"],
        "name": p["name"],
        "solvedCount": solved,
        "url": f"https://codeforces.com/problemset/problem/{p['contestId']}/{p['index']}",
    })

merged.sort(key=lambda x: (x["rating"], -x["solvedCount"], x["contestId"], x["index"]))

wb = Workbook()
ws = wb.active
ws.title = "Problems by Difficulty"

headers = ["Rating", "Contest ID", "Index", "Name", "Solved Count", "URL"]
header_font = Font(bold=True)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font

for row_idx, pb in enumerate(merged, 2):
    ws.cell(row=row_idx, column=1, value=pb["rating"])
    ws.cell(row=row_idx, column=2, value=pb["contestId"])
    ws.cell(row=row_idx, column=3, value=pb["index"])
    ws.cell(row=row_idx, column=4, value=pb["name"])
    ws.cell(row=row_idx, column=5, value=pb["solvedCount"])
    ws.cell(row=row_idx, column=6, value=pb["url"])

wb.save(OUTPUT_FILE)

print(f"Total problems (rated, PROGRAMMING): {len(merged)}")
print()

from collections import Counter
counts = Counter(pb["rating"] for pb in merged)
for rating in sorted(counts):
    print(f"  Rating {rating}: {counts[rating]} problems")

print()
print(f"Excel file saved to: {OUTPUT_FILE}")
